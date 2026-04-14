import os
import io
import pandas as pd
from datetime import datetime
from io import BytesIO
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, make_response, current_app
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from flask_apscheduler import APScheduler
from sqlalchemy import func, or_
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import mm
from fpdf import FPDF
from flask_mail import Mail, Message
from models import db, Usuario, Casa, Pago, Gasto, Configuracion, Deuda, IngresoExtra
from openpyxl.drawing.image import Image
import json
from openpyxl.styles import Font, Alignment
from xhtml2pdf import pisa
import locale

load_dotenv()

UMBRAL_ALERTA_CAJA = 30.00  # Se activa si hay menos de $30
app = Flask(__name__, static_folder='static', static_url_path='/static')
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')

# --- CONFIGURACIÓN DE CORREO ELECTRÓNICO ---
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'hernymejia@gmail.com' # Reemplaza con tu correo
app.config['MAIL_PASSWORD'] = 'wovytcycsqboesgl' # Reemplaza con tu contraseña de aplicación
app.config['MAIL_DEFAULT_SENDER'] = ('Aviso Conjunto Esperanza', 'hernymejia@gmail.com')

mail = Mail(app)

scheduler = APScheduler()


def tarea_cobro_mensual():
    with app.app_context():
        # 1. Obtener el valor de la alícuota desde la configuración
        config = Configuracion.query.filter_by(clave='valor_alicuota').first()
        valor = float(config.valor) if config else 20.0
        
        # 2. Sumar a todas las casas
        casas = Casa.query.all()
        for casa in casas:
            casa.saldo_total += valor
        
        db.session.commit()
        print(f"✅ Cobro automático ejecutado: ${valor} cargados a {len(casas)} casas.")

# Configurar la tarea para el día 1 de cada mes a las 00:00
@scheduler.task('cron', id='cobro_mensual', day=1, hour=0, minute=0)
def ejecutar_cron():
    tarea_cobro_mensual()

scheduler.init_app(app)
scheduler.start()


@app.route('/admin/casa/<int:id>')
@login_required
def detalle_casa(id):
    casa = Casa.query.get_or_404(id)
    # Importante: casa.deuda_anterior ya vive dentro del objeto casa
    deudas = Deuda.query.filter_by(casa_id=id).all()
    pagos = Pago.query.filter_by(casa_id=id).all()
    
    return render_template('admin/detalle_casa.html', 
                           casa=casa, 
                           deudas=deudas, 
                           pagos=pagos)


def registrar_pago_y_saldar_deuda(casa_id, monto_pagado, nota):
    # 1. Crear el registro del pago para el historial
    nuevo_pago = Pago(
        monto=monto_pagado,
        nota=nota,
        casa_id=casa_id,
        fecha=datetime.utcnow()
    )
    db.session.add(nuevo_pago)

    # 2. Buscar la deuda más antigua que no esté pagada
    deuda_pendiente = Deuda.query.filter_by(
        casa_id=casa_id, 
        pagado=False
    ).order_by(Deuda.anio.asc(), Deuda.mes.asc()).first()

    if deuda_pendiente:
        # 3. Si el monto coincide o es mayor, la marcamos como pagada
        # (Aquí podrías agregar lógica para abonos parciales si lo necesitas)
        if monto_pagado >= deuda_pendiente.monto:
            deuda_pendiente.pagado = True
            deuda_pendiente.fecha_pago = datetime.utcnow()
            mensaje = f"Pago registrado y deuda de {deuda_pendiente.nombre_mes()} saldada."
        else:
            mensaje = "Pago registrado, pero el monto es insuficiente para saldar la deuda completa."
    else:
        mensaje = "Pago registrado. No se encontraron deudas pendientes para esta casa."

    db.session.commit()
    return mensaje

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' # <--- ¡ESTO ES VITAL!

@login_manager.user_loader
def load_user(user_id):
    # Esto quita el error de LegacyAPIWarning
    return db.session.get(Usuario, int(user_id))

@app.context_processor
def inject_config():
    # Buscamos el número en la base de datos
    conf_ws = Configuracion.query.filter_by(clave='whatsapp_admin').first()
    # Si existe, lo convertimos a entero para limpiar puntos decimales, si no, uno por defecto
    num_ws = int(conf_ws.valor) if conf_ws else "593900000000"
    return dict(whatsapp_global=num_ws)



@app.route('/')
def index():
    # Esta función decide qué es lo primero que ve el usuario
    if current_user.is_authenticated:
        # Si ya inició sesión anteriormente
        return redirect(url_for('inicio'))
    else:
        # Si es un usuario nuevo o no se ha identificado
        return redirect(url_for('login'))

@app.route('/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    if request.method == 'POST':
        pass_actual = request.form.get('password_actual')
        nueva_pass = request.form.get('nueva_password')
        confirmar_pass = request.form.get('confirmar_password')

        # 1. Verificar si la contraseña actual es correcta
        if not check_password_hash(current_user.password, pass_actual):
            flash("❌ La contraseña actual es incorrecta.", "danger")
            return redirect(url_for('cambiar_password'))

        # 2. Verificar que las nuevas coincidan
        if nueva_pass != confirmar_pass:
            flash("❌ Las nuevas contraseñas no coinciden.", "danger")
            return redirect(url_for('cambiar_password'))

        # 3. Guardar la nueva contraseña (hasheada)
        current_user.password = generate_password_hash(nueva_pass)
        db.session.commit()
        
        flash("✅ Contraseña actualizada con éxito.",'success')
        return redirect(url_for('inicio'))

    return render_template('cambiar_password.html')

@app.route('/confirmar-pago-meses/<int:casa_id>', methods=['POST'])
@login_required
def confirmar_pago_meses(casa_id):
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    casa = db.get_or_404(Casa, casa_id)
    deuda_ids = request.form.getlist('deuda_ids')
    monto_abono = request.form.get('monto_abono') 
    metodo = request.form.get('metodo')
    
    # Variable para capturar el ID del pago y generar el recibo al final
    ultimo_pago_id = None

    try:
       
        if monto_abono:
            valor_abono = float(monto_abono)
            if valor_abono > 0:
                casa.deuda_anterior -= valor_abono 
                
                pago_abono = Pago(
                    monto=valor_abono,
                    casa_id=casa.id,
                    fecha=datetime.now(),
                    # CAMBIA ESTO:
                    nota=f"Pago de deuda anterior ({metodo})" 
                )
                db.session.add(pago_abono)
                db.session.flush()
                ultimo_pago_id = pago_abono.id
        # --- 2. PROCESAR MESES SELECCIONADOS (2026) ---
        for d_id in deuda_ids:
            deuda = db.session.get(Deuda, int(d_id))
            
            if deuda and not deuda.pagado:
                deuda.pagado = True
                deuda.fecha_pago = datetime.now()
                
                nuevo_pago = Pago(
                    monto=deuda.monto,
                    casa_id=casa.id,
                    deuda_id=deuda.id,
                    fecha=datetime.now(),
                    nota=f"Pago mes {deuda.mes}/{deuda.anio} ({metodo})"
                )
                db.session.add(nuevo_pago)
                db.session.flush()
                ultimo_pago_id = nuevo_pago.id

        # --- 3. RECALCULAR SALDO TOTAL FINAL ---
        nuevo_saldo_meses = db.session.query(func.sum(Deuda.monto)).filter(
            Deuda.casa_id == casa.id,
            Deuda.pagado == False
        ).scalar() or 0.0
        
        casa.saldo_total = float(nuevo_saldo_meses) + float(casa.deuda_anterior)

        db.session.commit()
        flash("✅ ¡Pago procesado con éxito!", "success")

        # Si hubo un pago (ya sea deuda anterior o mes), redirigir al recibo
        if ultimo_pago_id:
            return redirect(url_for('descargar_recibo', pago_id=ultimo_pago_id))

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error técnico: {str(e)}", "danger")

    return redirect(url_for('detalle_casa', id=casa_id))


@app.route('/inicio')
@login_required
def inicio():
    # --- 1. DATOS GLOBALES ---
    # Suma de alícuotas
    ingresos_alicuotas = db.session.query(func.sum(Pago.monto)).scalar() or 0.0
    
    # Suma de todos los ingresos extras
    ingresos_extras = db.session.query(func.sum(IngresoExtra.monto)).scalar() or 0.0
    
    # RECAUDACIÓN LOCAL: Comparación exacta con 'Alquiler Local'
    # Usamos func.trim para eliminar espacios accidentales antes o después del texto
    recaudacion_local = db.session.query(func.sum(IngresoExtra.monto)).filter(
        func.trim(IngresoExtra.categoria) == 'Alquiler Local'
    ).scalar() or 0.0
    
    # Cálculo de caja
    ingresos_totales = ingresos_alicuotas + ingresos_extras
    gastos_totales = db.session.query(func.sum(Gasto.monto)).scalar() or 0.0
    saldo_en_caja = ingresos_totales - gastos_totales

    # --- 2. VISTA PARA ADMINISTRADOR ---
    if current_user.rol == 'admin':
        # Estadísticas generales
        total_casas = Casa.query.count()
        casas_mora_count = Casa.query.filter(Casa.saldo_total > 0).count()
        casas_al_dia_count = total_casas - casas_mora_count
        datos_mora = [casas_al_dia_count, casas_mora_count]

        # Gastos por categoría (Dashboard)
        gastos_por_categoria = db.session.query(
            Gasto.categoria, func.sum(Gasto.monto)
        ).group_by(Gasto.categoria).all()

        labels_gastos = [cat if cat else "Sin Categoría" for cat, monto in gastos_por_categoria]
        valores_gastos = [float(monto) for cat, monto in gastos_por_categoria]

        # Actividad reciente
        ultimos_pagos_admin = Pago.query.order_by(Pago.id.desc()).limit(5).all()
        
        return render_template('admin/inicio_admin.html', 
                               saldo_caja=saldo_en_caja,
                               ingresos=ingresos_totales,
                               gastos=gastos_totales,
                               recaudacion_local=recaudacion_local,
                               total_casas=total_casas,
                               casas_mora=casas_mora_count,
                               labels_mora=json.dumps(['Al Día', 'En Mora']),
                               datos_mora=json.dumps(datos_mora),
                               labels_gastos=json.dumps(labels_gastos),
                               datos_gastos=json.dumps(valores_gastos),
                               pagos_recientes=ultimos_pagos_admin)

    # --- 3. VISTA PARA DUEÑO (Propietario) ---
    else:
        casa_dueno = Casa.query.filter_by(usuario_id=current_user.id).first()
        mis_pagos = []
        saldo_p = 0.0
        
        if casa_dueno:
            mis_pagos = Pago.query.filter_by(casa_id=casa_dueno.id).order_by(Pago.fecha.desc()).limit(5).all()
            saldo_p = casa_dueno.saldo_total or 0.0
        
        lista_morosos = Casa.query.filter(Casa.saldo_total > 0).order_by(Casa.saldo_total.desc()).all()
        ultimos_gastos = Gasto.query.order_by(Gasto.fecha.desc()).limit(5).all()
        
        return render_template('inicio_dueno.html', 
                               saldo_pendiente=saldo_p, 
                               saldo_caja=saldo_en_caja, 
                               recaudacion_local=recaudacion_local,
                               pagos=mis_pagos, 
                               gastos_comunidad=ultimos_gastos,
                               lista_morosos=lista_morosos)


@app.route('/descargar_recibo_extra/<int:id>')
@login_required
def descargar_recibo_extra(id):
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    ingreso = IngresoExtra.query.get_or_404(id)
    
    # Ruta al logo
    logo_path = os.path.join(current_app.root_path, 'static', 'img', 'logoEsperanza.png')
    
    # Si el archivo no existe físicamente, pasamos None para evitar errores
    if not os.path.exists(logo_path):
        logo_path = None
    
    html = render_template('admin/recibo_pdf.html', ingreso=ingreso, logo=logo_path)
    
    output = io.BytesIO()
    # Generamos el PDF
    pisa.CreatePDF(html.encode("UTF-8"), dest=output, encoding='UTF-8')
    
    output.seek(0)
    return make_response(output.read(), 200, {
        'Content-Type': 'application/pdf',
        'Content-Disposition': f'inline; filename=Recibo_{ingreso.id}.pdf'
    })


@app.route('/admin/configuracion', methods=['GET', 'POST'])
@login_required
def configuracion():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    from datetime import datetime # Importante tenerlo aquí o al inicio

    # 1. Buscar o crear configuraciones en la DB
    conf_alicuota = Configuracion.query.filter_by(clave='valor_alicuota').first()
    if not conf_alicuota:
        conf_alicuota = Configuracion(clave='valor_alicuota', valor=5.0)
        db.session.add(conf_alicuota)

    conf_whatsapp = Configuracion.query.filter_by(clave='whatsapp_admin').first()
    if not conf_whatsapp:
        conf_whatsapp = Configuracion(clave='whatsapp_admin', valor=593900000000) 
        db.session.add(conf_whatsapp)
    
    db.session.commit()

    # 2. Procesar cambios de valores si es POST
    if request.method == 'POST':
        nuevo_valor_ali = request.form.get('valor_alicuota')
        nuevo_ws = request.form.get('whatsapp_admin')
        
        if nuevo_valor_ali:
            conf_alicuota.valor = float(nuevo_valor_ali)
        if nuevo_ws:
            conf_whatsapp.valor = float(nuevo_ws)
            
        db.session.commit()
        flash("✅ Configuración actualizada con éxito", "success")
        return redirect(url_for('configuracion'))

    # 3. Lógica de meses dinámicos para el formulario de alícuotas
    anio_actual = datetime.now().year
    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    opciones_meses = []
    for anio in [anio_actual, anio_actual + 1]:
        for mes in meses_nombres:
            opciones_meses.append(f"{mes} {anio}")

    # 4. Enviamos todo al HTML
    return render_template('admin/configuracion.html', 
                           alicuota=conf_alicuota.valor, 
                           whatsapp=int(conf_whatsapp.valor),
                           opciones_meses=opciones_meses)

def generar_deudas_mensuales(mes, anio, monto_alicuota):
    """
    Crea un registro de deuda para cada casa que no tenga 
    una deuda registrada para el mes y año proporcionados.
    """
    casas = Casa.query.all()
    contador = 0
    
    for casa in casas:
        # Verificar si ya existe la deuda para evitar duplicados (por el UniqueConstraint)
        existe = Deuda.query.filter_by(casa_id=casa.id, mes=mes, anio=anio).first()
        
        if not existe:
            nueva_deuda = Deuda(
                casa_id=casa.id,
                mes=mes,
                anio=anio,
                monto=monto_alicuota,
                pagado=False
            )
            db.session.add(nueva_deuda)
            contador += 1
    
    db.session.commit()
    print(f"Se generaron {contador} registros de deuda para el periodo {mes}/{anio}.")


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = Usuario.query.filter_by(username=username).first()
        
        # Esta es la parte clave:
        if user and check_password_hash(user.password, password):
            login_user(user)
            flash('¡Bienvenido de nuevo!','success')
            return redirect(url_for('inicio'))
        else:
            flash('Usuario o contraseña incorrectos','success')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    from flask_login import logout_user
    logout_user() # Borra la sesión del navegador
    flash("Has cerrado sesión correctamente.")
    return redirect(url_for('login'))


@app.route('/admin/registrar-pago/<int:casa_id>', methods=['GET'])
@login_required
def registrar_pago(casa_id):
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    # Buscamos la casa específica por su ID
    casa = db.session.get(Casa, casa_id)
    
    if not casa:
        flash("Casa no encontrada", "danger")
        return redirect(url_for('lista_casas'))

    # Renderizamos el template pasando el objeto 'casa' que el HTML necesita
    return render_template('admin/registrar_pago.html', casa=casa)

@app.route('/admin/descargar-reporte-deudas')
@login_required
def descargar_reporte_deudas():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    # 1. Obtener todas las casas
    casas = Casa.query.order_by(Casa.numero_casa).all()

    # 2. Crear una lista de diccionarios con los datos
    datos = []
    for c in casas:
        datos.append({
            "Número de Casa": c.numero_casa,
            "Propietario": c.dueno_nombre,
            "Deuda Pendiente ($)": c.saldo_total,
            "Estado": "Al día" if c.saldo_total <= 0 else "Con Deuda"
        })

    # 3. Convertir a DataFrame de Pandas
    df = pd.DataFrame(datos)

    # 4. Guardar el Excel en memoria (sin crear archivos físicos en el servidor)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte Deudas')
    
    output.seek(0)

    # 5. Enviar el archivo al usuario
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Reporte_Deudas_SISE.xlsx"
    )


@app.route('/admin/iniciar-nuevo-anio', methods=['POST'])
@login_required
def iniciar_nuevo_anio():
    from datetime import datetime
    anio_nuevo = datetime.now().year
    
    # En este caso, como quieres mantener el saldo acumulado, 
    # NO reseteamos casa.saldo_total a cero.
    # Solo registramos en el log o enviamos el mensaje de éxito.
    
    try:
        # Aquí podrías opcionalmente crear una nota en el historial 
        # indicando que se inició el ciclo fiscal {anio_nuevo}
        db.session.commit()
        flash(f"✅ Ciclo {anio_nuevo} iniciado. Los saldos pendientes se han mantenido para el cobro actual.", "success")
    except Exception as e:
        flash(f"Error al procesar: {str(e)}", "danger")
        
    return redirect(url_for('vista_reportes'))


@app.route('/admin/recibo/<int:pago_id>')
@login_required
def descargar_recibo(pago_id):
    pago = Pago.query.get_or_404(pago_id)
    casa = Casa.query.get(pago.casa_id)
    
    # Cálculo del saldo pendiente (Alícuotas 2026 + Deuda Anterior)
    saldo_pendiente_meses = db.session.query(db.func.sum(Deuda.monto)).filter(
        Deuda.casa_id == casa.id,
        Deuda.pagado == False
    ).scalar() or 0.0
    total_pendiente = float(saldo_pendiente_meses) + float(casa.deuda_anterior or 0)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    
    def dibujar_recibo(c, tipo_copia):
        # --- LOGO Y ENCABEZADO ---
        logo_path = os.path.join(app.root_path, 'static', 'img', 'logoEsperanza.png')
        if os.path.exists(logo_path):
            c.drawImage(logo_path, 15*mm, 115*mm, width=20*mm, height=20*mm, preserveAspectRatio=True)

        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.darkblue)
        c.drawString(40*mm, 125*mm, "SISEsperanza")
        
        c.setFont("Helvetica-Bold", 9)
        c.setFillColor(colors.grey)
        c.drawRightString(195*mm, 125*mm, tipo_copia)

        # --- DATOS BÁSICOS ---
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(15*mm, 105*mm, f"RECIBO N°: {pago.id:06d}")
        c.drawRightString(195*mm, 105*mm, f"FECHA: {pago.fecha.strftime('%d/%m/%Y')}")

        # --- CUADRO DEL SOCIO ---
        c.rect(15*mm, 85*mm, 180*mm, 15*mm, stroke=1, fill=0)
        c.setFont("Helvetica", 9)
        c.drawString(18*mm, 95*mm, f"PROPIETARIO: {casa.dueno_nombre}")
        c.drawString(18*mm, 88*mm, f"CASA / VILLA: {casa.numero_casa}")

        # --- CONCEPTO DINÁMICO (AQUÍ ESTÁ LA CORRECCIÓN) ---
        c.setFont("Helvetica-Bold", 10)
        # Si el pago tiene nota, la usa. Si no, pone el default.
        texto_concepto = pago.nota if pago.nota else "Pago de alícuota mensual"
        c.drawString(15*mm, 75*mm, f"CONCEPTO: {texto_concepto}")

        # --- CUADRO DE TOTAL ---
        c.setFillColor(colors.black)
        c.rect(145*mm, 68*mm, 50*mm, 10*mm, fill=1)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(148*mm, 71*mm, "PAGADO:")
        c.drawRightString(192*mm, 71*mm, f"${'{:,.2f}'.format(pago.monto)}")

        # --- SALDO PENDIENTE ---
        c.setFillColor(colors.black)
        if total_pendiente > 0:
            c.setFont("Helvetica-Bold", 10)
            c.drawString(15*mm, 60*mm, f"SALDO PENDIENTE TOTAL: ${'{:,.2f}'.format(total_pendiente)}")
        else:
            c.setFont("Helvetica-Bold", 10)
            c.drawString(15*mm, 60*mm, "ESTADO DE CUENTA: AL DÍA")

        # --- FIRMA ---
        c.line(70*mm, 25*mm, 140*mm, 25*mm)
        c.setFont("Helvetica-Oblique", 8)
        c.drawCentredString(105*mm, 20*mm, "Firma Autorizada")

    # Dibujar las dos copias en la misma hoja
    p.saveState()
    p.translate(0, 148.5*mm) 
    dibujar_recibo(p, "ORIGINAL - SOCIO")
    p.restoreState()

    p.setDash([3, 6])
    p.line(0, 148.5*mm, 210*mm, 148.5*mm)
    p.setDash([])

    dibujar_recibo(p, "COPIA - ADMINISTRACIÓN")
    
    p.showPage()
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Recibo_{pago.id}.pdf", mimetype='application/pdf')

def generar_recibo_pdf(pago, casa):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # --- Cálculo de Saldo Pendiente ---
    saldo_restante = db.session.query(db.func.sum(Deuda.monto)).filter(
        Deuda.casa_id == casa.id,
        Deuda.pagado == False
    ).scalar() or 0.0
    total_pendiente = float(saldo_restante) + float(casa.deuda_anterior or 0)

    # --- Diseño del Recibo ---
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, "🏘️ SiSEsperanza - Comprobante de Pago")
    
    p.setLineWidth(1)
    p.line(100, 740, 500, 740)
    
    p.setFont("Helvetica", 12)
    p.drawString(100, 710, f"Recibo N°: {pago.id:06d}")
    p.drawString(100, 690, f"Fecha: {pago.fecha.strftime('%d/%m/%Y')}")
    
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, 650, "DATOS DEL PROPIETARIO:")
    p.setFont("Helvetica", 12)
    p.drawString(100, 630, f"Nombre: {casa.dueno_nombre}")
    p.drawString(100, 610, f"Casa N°: {casa.numero_casa}")
    
    # --- DETALLE DEL PAGO (CORREGIDO) ---
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, 570, "DETALLE DEL PAGO:")
    p.setFont("Helvetica", 12)
    
    # Aquí está el truco: usamos pago.nota que viene de la DB
    texto_concepto = pago.nota if pago.nota else "Pago de alícuota mensual"
    p.drawString(100, 550, f"Concepto: {texto_concepto}")
    
    p.setFont("Helvetica-Bold", 14)
    p.setFillColorRGB(0, 0.5, 0) 
    p.drawString(100, 520, f"MONTO PAGADO: ${'{:,.2f}'.format(pago.monto)}")
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica-Bold", 11)
    p.drawString(100, 490, f"SALDO PENDIENTE TOTAL: ${'{:,.2f}'.format(total_pendiente)}")

    p.line(100, 470, 500, 470)
    p.setFont("Helvetica-Oblique", 10)
    p.drawString(100, 450, "Este documento sirve como soporte legal de su pago.")
    p.drawString(100, 435, "Gracias por su contribución al mantenimiento del conjunto.")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    return buffer

@app.route('/admin/registrar-casa', methods=['GET', 'POST'])
@login_required
def registrar_casa():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        try:
            num = request.form.get('numero_casa')
            nombre = request.form.get('dueno_nombre')
            # Capturamos la deuda de años anteriores (2025 hacia atrás)
            deuda_inicial = float(request.form.get('deuda_anterior') or 0.0)

            # Verificamos si la casa ya existe
            existe = Casa.query.filter_by(numero_casa=num).first()
            if existe:
                flash(f"⚠️ La casa {num} ya está registrada.", "warning")
                return redirect(url_for('registrar_casa'))

            # Creamos la nueva casa
            nueva_casa = Casa(
                numero_casa=num,
                dueno_nombre=nombre,
                deuda_anterior=deuda_inicial, # Saldo persistente de 2025
                saldo_total=deuda_inicial      # El total inicial (se irá sumando con los meses 2026)
            )

            db.session.add(nueva_casa)
            db.session.commit()
            
            flash(f"✅ Casa {num} registrada con éxito.", "success")
            return redirect(url_for('inicio'))

        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error al registrar: {str(e)}", "danger")

    return render_template('admin/registrar_casa.html')

@app.route('/admin/generar-mensualidad', methods=['POST'])
@login_required
def generar_alicuotas(): 
    meses_map = {
        "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4, 
        "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8, 
        "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
    }

    seleccion = request.form.get('mes')
    
    # OPCIONAL PERO RECOMENDADO: Traer el monto desde la base de datos
    conf = Configuracion.query.filter_by(clave='valor_alicuota').first()
    monto_cuota = conf.valor if conf else 20.0
    
    if not seleccion:
        flash("Seleccione un mes válido.", "danger")
        return redirect(url_for('configuracion'))

    try:
        nombre_mes, anio_valor = seleccion.split()
        mes_numero = meses_map[nombre_mes]
        anio_numero = int(anio_valor)
    except:
        flash("Error en el formato del mes.", "danger")
        return redirect(url_for('configuracion'))

    # Candado para no duplicar deudas del mismo mes
    existe = Deuda.query.filter_by(mes=mes_numero, anio=anio_numero).first()
    if existe:
        flash(f"⚠️ Las alícuotas para {seleccion} ya existen.", "warning")
        return redirect(url_for('configuracion'))

    casas = Casa.query.all()
    for casa in casas:
        nueva_deuda = Deuda(
            monto=monto_cuota,
            mes=mes_numero,
            anio=anio_numero,
            pagado=False,
            casa_id=casa.id
        )
        db.session.add(nueva_deuda)
        
        if casa.saldo_total is None:
            casa.saldo_total = 0.0
    
      
        saldo_meses_pendientes = db.session.query(db.func.sum(Deuda.monto)).filter(
            Deuda.casa_id == casa.id, Deuda.pagado == False
        ).scalar() or 0.0
        
        casa.saldo_total = float(saldo_meses_pendientes) + float(casa.deuda_anterior or 0)

    db.session.commit()
    flash(f"✅ Alícuotas de {seleccion} generadas por ${monto_cuota}", "success")
    return redirect(url_for('configuracion'))

@app.route('/admin/buscar-pagos')
@login_required
def buscar_pagos():
    query = request.args.get('q', '')
    
    # Buscamos el pago relacionado para poder descargar el recibo
    if query:
        # Unimos Pago con Casa para la búsqueda
        resultados = db.session.query(Pago, Casa).join(Casa).filter(
            db.or_(
                Casa.dueno_nombre.ilike(f'%{query}%'),
                Casa.numero_casa.ilike(f'%{query}%'),
                Pago.nota.ilike(f'%{query}%')
            )
        ).order_by(Pago.fecha.desc()).all()
    else:
        resultados = db.session.query(Pago, Casa).join(Casa).order_by(Pago.fecha.desc()).limit(20).all()
    
    return render_template('admin/buscar_pagos.html', resultados=resultados, query=query)

@app.route('/admin/crear-dueno', methods=['GET', 'POST'])
@login_required
def crear_dueno():
    if current_user.rol != 'admin':
        return "Acceso denegado", 403

    if request.method == 'POST':
        numero_casa = request.form.get('numero_casa')
        nombre_dueno = request.form.get('nombre')
        deuda_inicial = request.form.get('saldo_total', 0)
        username = request.form.get('username')
        password = request.form.get('password')

        # 1. Crear la Casa
        nueva_casa = Casa(
            numero_casa=numero_casa, 
            dueno_nombre=nombre_dueno, 
            saldo_total=float(deuda_inicial)
        )
        db.session.add(nueva_casa)
        db.session.flush() # Esto genera el ID de la casa antes de guardar

        # 2. Crear el Usuario para ese dueño
        nuevo_usuario = Usuario(
            username=username,
            password=password, # Idealmente usaría bcrypt después
            rol='dueno',
            casa_id=nueva_casa.id
        )
        db.session.add(nuevo_usuario)
        db.session.commit()
        
        flash(f'Casa {numero_casa} y usuario {username} creados con éxito')
        return redirect(url_for('registrar_pago'))

    return render_template('admin/crear_dueno.html')

@app.route('/casas')
@login_required
def lista_casas():
    # Buscamos todas las casas
    todas_las_casas = Casa.query.order_by(Casa.numero_casa.asc()).all()
    # IMPORTANTE: Apuntar a la subcarpeta 'admin/'
    return render_template('admin/lista_casas.html', casas=todas_las_casas)



@app.route('/admin/propietarios')
@login_required
def lista_propietarios():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    # Traemos a todos los usuarios que tienen el rol de 'usuario'
    # .all() es fundamental para obtener la lista
    lista_users = Usuario.query.filter_by(rol='propietario').all()
    todos = Usuario.query.all()
    print(f"Total en DB: {len(todos)} | Roles encontrados: {[u.rol for u in todos]}")
    return render_template('propietarios.html', propietarios=lista_users)
 


@app.route('/admin/reportes/morosos')
@login_required
def reporte_morosos():
    # Buscamos solo casas que tengan deudas pendientes
    morosos = Casa.query.join(Deuda).filter(Deuda.pagado == False).distinct().all()
    
    # Suma total de lo que deben todos los morosos
    total_deuda = db.session.query(db.func.sum(Deuda.monto)).filter(Deuda.pagado == False).scalar() or 0
    
    return render_template('admin/reporte_morosos.html', casas=morosos, total_deuda=total_deuda)

@app.route('/admin/registrar-usuario', methods=['GET', 'POST'])
@login_required
def registrar_usuario():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        # 1. Capturar datos del formulario (el name del HTML debe ser 'correo')
        username = request.form.get('username')
        password = request.form.get('password')
        cedula = request.form.get('cedula')
        telefono = request.form.get('telefono')
        correo_form = request.form.get('correo') 
        casa_id = request.form.get('casa_id') 

        # 2. Crear el nuevo objeto Usuario
        nuevo_usuario = Usuario(
            username=username,
            password=generate_password_hash(password),
            cedula=cedula,
            telefono=telefono,
            correo=correo_form,  # <--- SE GUARDA AQUÍ
            rol='propietario'
        )
        
        db.session.add(nuevo_usuario)
        db.session.flush() # Genera el ID del usuario antes del commit

        # 3. Vincular con la casa si se seleccionó una
        if casa_id:
            casa = db.session.get(Casa, int(casa_id))
            if casa:
                casa.usuario_id = nuevo_usuario.id
                casa.dueno_nombre = username 
        
        try:
            db.session.commit()
            flash("✅ Propietario creado y correo guardado correctamente.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar: {str(e)}", "danger")
            
        return redirect(url_for('lista_propietarios'))

    casas_disponibles = Casa.query.filter_by(usuario_id=None).all()
    return render_template('admin/registrar_usuario.html', casas=casas_disponibles)


@app.route('/admin/registrar-gasto', methods=['GET', 'POST'])
@login_required
def registrar_gasto():
    if current_user.rol != 'admin':
        return redirect(url_for('ver_estado_cuenta'))

    if request.method == 'POST':
        desc = request.form.get('descripcion')
        monto = request.form.get('monto')
        cat = request.form.get('categoria')
        recibo = request.form.get('numero_recibo') # <--- NUEVO: Captura el recibo

        nuevo_gasto = Gasto(
            descripcion=desc,
            monto=float(monto),
            categoria=cat,
            numero_recibo=recibo # <--- NUEVO: Asegúrate que tu clase Gasto tenga este campo
        )
        db.session.add(nuevo_gasto)
        db.session.commit()
        flash("✅ Gasto registrado con éxito")
        return redirect(url_for('registrar_gasto'))

    gastos = Gasto.query.order_by(Gasto.fecha.desc()).all()
    total_gastos = sum(g.monto for g in gastos)
    return render_template('admin/registrar_gasto.html', gastos=gastos, total_gastos=total_gastos)


@app.route('/admin/descargar-pdf-gastos')
@login_required
def descargar_pdf_gastos():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    # 1. Obtener filtros de la URL (si existen)
    hoy = datetime.now()
    mes_filtro = request.args.get('mes', hoy.month, type=int)
    anio_filtro = request.args.get('anio', hoy.year, type=int)

    # 2. Filtrar los gastos en la base de datos
    gastos = Gasto.query.filter(
        db.extract('month', Gasto.fecha) == mes_filtro,
        db.extract('year', Gasto.fecha) == anio_filtro
    ).order_by(Gasto.fecha.desc()).all()
    
    # 3. Configurar el PDF (FPDF)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    
    # Título dinámico
    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes = meses_nombres[mes_filtro - 1]

    pdf.cell(190, 10, "SISTEMA SISESPERANZA", ln=True, align='C')
    pdf.set_font("Arial", '', 12)
    pdf.cell(190, 10, f"Informe de Gastos: {nombre_mes} {anio_filtro}", ln=True, align='C')
    pdf.ln(10)

    # ... (Resto de tu código de la tabla del PDF igual que antes) ...

    # Guardar y enviar
    output_path = os.path.join('instance', f'informe_gastos_{nombre_mes}_{anio_filtro}.pdf')
    pdf.output(output_path)
    return send_file(output_path, as_attachment=True)


@app.route('/admin/editar-propietario/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_propietario(id):
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))
    
    usuario = db.session.get(Usuario, id)
    if not usuario:
        flash("Usuario no encontrado", "warning")
        return redirect(url_for('lista_propietarios'))

    if request.method == 'POST':
        # 1. Actualizar datos básicos
        usuario.username = request.form.get('username')
        usuario.cedula = request.form.get('cedula')
        usuario.telefono = request.form.get('telefono')
        usuario.correo = request.form.get('correo') # <--- ACTUALIZA EL CORREO
        
        # 2. Manejo de contraseña (solo si se escribe una nueva)
        nueva_pass = request.form.get('password')
        if nueva_pass and nueva_pass.strip() != "":
            usuario.password = generate_password_hash(nueva_pass)
            
        # 3. Lógica de vinculación de casa
        nueva_casa_id = request.form.get('casa_id')
        
        # Desvincular de la casa anterior
        casa_anterior = Casa.query.filter_by(usuario_id=usuario.id).first()
        if casa_anterior:
            casa_anterior.usuario_id = None
        
        # Vincular a la nueva casa
        if nueva_casa_id:
            nueva_casa = db.session.get(Casa, int(nueva_casa_id))
            if nueva_casa:
                nueva_casa.usuario_id = usuario.id
                nueva_casa.dueno_nombre = usuario.username 
            
        try:
            db.session.commit()
            flash("✅ Datos y correo actualizados correctamente", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al actualizar: {str(e)}", "danger")

        return redirect(url_for('lista_propietarios'))

    # Para el formulario: mostrar casas libres o la que ya tiene el usuario
    casas_disponibles = Casa.query.filter((Casa.usuario_id == None) | (Casa.usuario_id == usuario.id)).all()
    return render_template('admin/editar_propietario.html', usuario=usuario, casas=casas_disponibles)

@app.route('/propietario/descargar-mi-estado-pdf')
@login_required
def descargar_mi_estado_pdf():
    # 1. Buscar la casa vinculada al usuario logueado
    casa = Casa.query.filter_by(usuario_id=current_user.id).first()
    
    if not casa:
        flash("No tienes una propiedad asignada para generar el reporte.")
        return redirect(url_for('mi_estado'))

    # 2. Configurar el PDF
    pdf = FPDF()
    pdf.add_page()
    
    # Encabezado con Estilo
    pdf.set_font("Arial", 'B', 16)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(190, 10, "ESTADO DE CUENTA - SISEsperanza", ln=True, align='C')
    
    pdf.set_font("Arial", '', 10)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(190, 10, f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='C')
    pdf.ln(10)

    # Datos de la Propiedad
    pdf.set_fill_color(241, 245, 249)
    pdf.set_font("Arial", 'B', 12)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(190, 10, f" PROPIETARIO: {casa.dueno_nombre}", 1, 1, 'L', True)
    pdf.cell(95, 10, f" CASA N°: {casa.numero_casa}", 1, 0, 'L')
    pdf.cell(95, 10, f" DEUDA ACTUAL: ${casa.saldo_total:,.2f}", 1, 1, 'L')
    pdf.ln(10)

    # Tabla de Pagos Realizados
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(190, 10, "HISTORIAL DE PAGOS REGISTRADOS", ln=True)
    
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(40, 10, " Fecha", 1, 0, 'C', True)
    pdf.cell(100, 10, " Concepto", 1, 0, 'C', True)
    pdf.cell(50, 10, " Monto", 1, 1, 'C', True)

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", '', 10)
    
    for pago in casa.pagos:
        pdf.cell(40, 10, pago.fecha.strftime('%d/%m/%Y'), 1, 0, 'C')
        pdf.cell(100, 10, f" {pago.concepto or 'Pago de Alícuota'}", 1, 0, 'L')
        pdf.cell(50, 10, f"${pago.monto:,.2f}", 1, 1, 'R')

    # Pie de página informativo
    pdf.ln(10)
    pdf.set_font("Arial", 'I', 9)
    pdf.set_text_color(100, 116, 139)
    pdf.multi_cell(190, 5, "Nota: Este documento es un reporte informativo de sus aportaciones. Si encuentra alguna inconsistencia, por favor contacte a la administración con su comprobante físico.", align='C')

    # Enviar archivo
    output_path = os.path.join('instance', f'Estado_Cuenta_Casa_{casa.numero_casa}.pdf')
    pdf.output(output_path)
    return send_file(output_path, as_attachment=True)


@app.route('/admin/enviar-avisos-correo')
@login_required
def enviar_avisos_correo():
    casas_con_deuda = Casa.query.filter(Casa.saldo_total > 0).all()
    print(f"DEBUG: Se encontraron {len(casas_con_deuda)} casas con deuda.") # <--- MIRA ESTO
    
    contador_exitos = 0
    for casa in casas_con_deuda:
        usuario = Usuario.query.filter_by(id=casa.usuario_id).first()
        
        if usuario and usuario.correo:
            msg = Message(
                subject=f"Estado de Cuenta Pendiente - Casa {casa.numero_casa}",
                recipients=[usuario.correo]
            )
            msg.body = f"""
            Estimado(a) {usuario.username},

            Le informamos que su propiedad (Casa {casa.numero_casa}) mantiene un saldo pendiente 
            de ${casa.saldo_total:,.2f} en la administración de SISE Esperanza.

            Le solicitamos cordialmente realizar su pago para el mantenimiento de la urbanización.

            Atentamente,
            La Administración.
            """
            try:
                mail.send(msg)
                contador_exitos += 1
            except Exception as e:
                print(f"Error enviando correo a {usuario.correo}: {e}")

    flash(f"✅ Proceso finalizado. Se enviaron {contador_exitos} correos de recordatorio.", "success")
    return redirect(url_for('inicio')) # O a la vista de reportes


@app.route('/admin/pagos-globales')
@login_required
def pagos_globales():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))
    
    # Capturamos el término de búsqueda del input 'q'
    query = request.args.get('q', '')

    if query:
        # Filtramos por número de casa, nombre del dueño o nota del pago
        pagos = db.session.query(Pago).join(Casa).filter(
            or_(
                Casa.numero_casa.ilike(f'%{query}%'),
                Casa.dueno_nombre.ilike(f'%{query}%'),
                Pago.nota.ilike(f'%{query}%')
            )
        ).order_by(Pago.fecha.desc()).all()
    else:
        # Si no hay búsqueda, mostramos todos los pagos
        pagos = Pago.query.order_by(Pago.fecha.desc()).all()

    total_recaudado = sum(p.monto for p in pagos)
    
    return render_template('admin/pagos_globales.html', 
                           pagos=pagos, 
                           total=total_recaudado, 
                           busqueda=query)


@app.route('/admin/editar-casa/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_casa(id):
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    casa = db.session.get(Casa, id)
    
    if not casa:
        flash("❌ Casa no encontrada", "danger")
        return redirect(url_for('lista_casas'))

    if request.method == 'POST':
        try:
            # 1. Actualizamos solo los campos existentes
            casa.numero_casa = request.form.get('numero_casa')
            casa.dueno_nombre = request.form.get('dueno_nombre')
            
            # 2. Capturamos la deuda anterior
            deuda_ant_nueva = float(request.form.get('deuda_anterior') or 0)
            casa.deuda_anterior = deuda_ant_nueva
            
            # 3. Recalculamos el saldo total (Meses 2026 + Deuda 2025)
            meses_pendientes = db.session.query(db.func.sum(Deuda.monto)).filter(
                Deuda.casa_id == casa.id,
                Deuda.pagado == False
            ).scalar() or 0.0
            
            casa.saldo_total = float(meses_pendientes) + float(casa.deuda_anterior)

            db.session.commit()
            flash(f"✅ Casa {casa.numero_casa} actualizada correctamente.", "success")
            return redirect(url_for('lista_casas'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error al actualizar: {str(e)}", "danger")

    return render_template('admin/editar_casa.html', casa=casa)


@app.route('/admin/eliminar-casa/<int:id>', methods=['POST'])
@login_required
def eliminar_casa(id):
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))
    
    casa = Casa.query.get_or_404(id)
    try:
        # Nota: Si la casa tiene pagos o deudas asociadas, 
        # esto podría dar error dependiendo de cómo definiste los modelos.
        db.session.delete(casa)
        db.session.commit()
        flash(f"🗑️ Casa {casa.numero_casa} eliminada con éxito", "warning")
    except Exception as e:
        db.session.rollback()
        flash("No se puede eliminar la casa porque tiene deudas o pagos registrados.", "danger")
        
    return redirect(url_for('lista_casas'))


@app.route('/admin/eliminar-propietario/<int:id>')
@login_required
def eliminar_propietario(id):
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))
    
    usuario = Usuario.query.get_or_404(id)
    try:
        db.session.delete(usuario)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return f"No se puede eliminar: el usuario tiene registros asociados. {e}"
        
    return redirect(url_for('lista_propietarios'))


@app.route('/inicio_admin') # Asegúrate de que el nombre coincida con tus plantillas
@login_required
def inicio_admin():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))
    
    # ... (todo tu código actual de la función inicio) ...
    # Pero asegúrate de pasar 'casas' para que la tabla de abajo no salga vacía
    todas_las_casas = Casa.query.order_by(Casa.numero_casa.asc()).all()
    
    return render_template('admin/inicio_admin.html', 
                           casas=todas_las_casas, # <--- IMPORTANTE
                           ingresos=total_ingresos, 
                           egresos=total_egresos, 
                           balance=balance,
                           movimientos=movimientos)


@app.route('/admin/registrar-pago-mes/<int:deuda_id>', methods=['POST'])
@login_required
def registrar_pago_mes(deuda_id):
    deuda = Deuda.query.get_or_404(deuda_id)
    casa = Casa.query.get(deuda.casa_id) # Obtenemos la casa para actualizar su saldo
    
    if not deuda.pagado:
        # 1. Marcar la mensualidad como pagada
        deuda.pagado = True
        deuda.fecha_pago = datetime.now()
        
        # 2. RESTAR de la deuda global de la casa (AQUÍ ESTABA EL FALLO)
        # Restamos el monto de la deuda actual de la casa
        casa.saldo_total = float(casa.saldo_total) - float(deuda.monto)
        
        # 3. Crear el registro del movimiento (Ingreso)
        nuevo_pago = Pago(
            monto=deuda.monto,
            casa_id=deuda.casa_id,
            deuda_id=deuda.id,
            fecha=datetime.now()
            # nota="Pago de mensualidad" (Usa el nombre de columna que definimos antes)
        )
        
        try:
            db.session.add(nuevo_pago)
            db.session.commit()
            flash(f"✅ Pago registrado. Saldo de casa actualizado.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error al procesar: {str(e)}", "danger")
    else:
        flash("Esta deuda ya figuraba como pagada.", "warning")
        
    return redirect(url_for('detalle_casa', id=deuda.casa_id))

@app.route('/admin/recalcular-saldos')
@login_required
def recalcular_saldos():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    try:
        casas = Casa.query.all()
        for casa in casas:
            # Sumamos todas las deudas pendientes de esta casa
            total_pendiente = db.session.query(db.func.sum(Deuda.monto)).filter(
                Deuda.casa_id == casa.id,
                Deuda.pagado == False
            ).scalar() or 0.0
            
            # Actualizamos el saldo de la casa
            casa.saldo_total = float(total_pendiente)
        
        db.session.commit()
        flash("✅ ¡Saldos sincronizados! Todos los totales coinciden ahora con las deudas pendientes.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al sincronizar: {str(e)}", "danger")

    return redirect(url_for('lista_casas'))

@app.route('/mi-cuenta')
@login_required
def mi_cuenta():
    # 1. Buscamos la casa que le pertenece al usuario logueado
    # Usamos dueno_nombre porque así se llama en tu modelo Casa
    casa = Casa.query.filter_by(dueno_nombre=current_user.nombre).first()

    if not casa:
        # Si no lo encuentra por nombre, el administrador debe verificar 
        # que el nombre del usuario y el 'dueno_nombre' sean idénticos.
        flash("No se encontró una propiedad vinculada a su cuenta.")
        return redirect(url_for('inicio'))

    # 2. Obtenemos los pagos usando la relación que definiste: casa.pagos
    # Esto busca automáticamente en la tabla Pagos todos los que tengan el casa_id de esta casa
    mis_pagos = casa.pagos

    return render_template('estado_cuenta.html', casa=casa, pagos=casa.pagos)

@app.route('/admin/dashboard')
@login_required
def dashboard():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    # 1. Datos para gráfico de Ingresos vs Gastos
    total_ingresos = db.session.query(func.sum(Pago.monto)).scalar() or 0
    total_gastos = db.session.query(func.sum(Gasto.monto)).scalar() or 0
    
    # 2. Datos de Mora vs Recaudación
    total_mora = db.session.query(func.sum(Casa.saldo_total)).filter(Casa.saldo_total > 0).scalar() or 0
    
    # 3. Datos para gráfico de barras (Últimos 6 meses)
    # Esto requiere una consulta un poco más compleja para agrupar por mes
    # Por ahora, enviamos los totales principales
    
    return render_template('admin/dashboard.html', 
                           ingresos=total_ingresos, 
                           gastos=total_gastos, 
                           mora=total_mora,
                           saldo=total_ingresos - total_gastos)


@app.route('/admin/reporte')
@login_required
def reporte_general():
    if current_user.rol != 'admin':
        return "Acceso denegado", 403
    
    # Suma total de todos los pagos registrados en la historia
    total_recaudado = db.session.query(func.sum(Pago.monto)).scalar() or 0
    # Lista de todas las casas con sus respectivos saldos
    casas = Casa.query.all()
    
    return render_template('admin/reporte.html', total=total_recaudado, casas=casas)


@app.route('/admin/otros-ingresos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_ingreso_extra():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        try:
            monto = float(request.form.get('monto'))
            concepto = request.form.get('concepto')
            categoria = request.form.get('categoria')
            fecha_str = request.form.get('fecha')
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d')

            # USAMOS EL NUEVO MODELO IngresoExtra 1
            nuevo = IngresoExtra(
                monto=monto,
                concepto=concepto,
                categoria=categoria,
                fecha=fecha
            )

            db.session.add(nuevo)
            db.session.commit()
            flash(f"✅ {categoria} registrado correctamente", "success")
            return redirect(url_for('lista_otros_ingresos'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error al registrar: {str(e)}", "danger")

    return render_template('admin/otros_ingresos_form.html', hoy=datetime.now().strftime('%Y-%m-%d'))

@app.route('/admin/reportes')
@login_required
def vista_reportes():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    # --- PASO EXTRA: Sincronizar saldos antes de mostrar el reporte ---
    casas_db = Casa.query.all()
    for casa in casas_db:
        # Sumamos lo que realmente debe en la tabla Deuda
        total_deuda = db.session.query(db.func.sum(Deuda.monto)).filter(
            Deuda.casa_id == casa.id,
            Deuda.pagado == False
        ).scalar() or 0.0
        casa.saldo_total = float(total_deuda)
    
    db.session.commit() # Guardamos los saldos reales
    # ---------------------------------------------------------------

    # Ahora sí, pedimos los datos para el reporte
    ingresos_db = db.session.query(func.sum(Pago.monto)).scalar() or 0.0
    gastos_db = db.session.query(func.sum(Gasto.monto)).scalar() or 0.0
    deuda_db = db.session.query(func.sum(Casa.saldo_total)).scalar() or 0.0
    
    saldo_en_caja = ingresos_db - gastos_db

    # Traemos las casas que tienen deuda > 0 para la tabla
    casas_con_mora = Casa.query.filter(Casa.saldo_total > 0).order_by(Casa.numero_casa).all()

    return render_template('admin/reportes.html', 
                           ingresos=ingresos_db, 
                           gastos=gastos_db, 
                           saldo=saldo_en_caja,
                           deuda_pendiente=deuda_db,
                           casas_con_deuda=casas_con_mora) # <--- Asegúrate que este nombre coincida con el HTML


@app.route('/admin/otros-ingresos')
@login_required
def lista_otros_ingresos():
    ingresos = IngresoExtra.query.order_by(IngresoExtra.fecha.desc()).all()
    return render_template('admin/otros_ingresos_lista.html', ingresos=ingresos)


@app.route('/admin/otros-ingresos/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_ingreso_extra(id):
    ingreso = IngresoExtra.query.get_or_404(id)
    db.session.delete(ingreso)
    db.session.commit()
    flash("🗑️ Registro eliminado", "warning")
    return redirect(url_for('lista_otros_ingresos'))

@app.route('/admin/reporte-anual-excel')
@login_required
def reporte_anual_excel():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    casas = Casa.query.order_by(Casa.numero_casa).all()
    anio_actual = datetime.now().year
    
    datos = []
    for c in casas:
        # Calculamos cuánto pagó en el año
        total_pagado_anio = db.session.query(func.sum(Pago.monto)).filter(
            Pago.casa_id == c.id,
            func.extract('year', Pago.fecha) == anio_actual
        ).scalar() or 0.0

        datos.append({
            "Casa": c.numero_casa,
            "Propietario": c.dueno_nombre,
            "Deuda Inicial (Ene 1)": c.deuda_anterior,
            "Total Pagado en el Año": total_pagado_anio,
            "Saldo Pendiente Actual": c.saldo_total,
            "Estado": "AL DÍA" if c.saldo_total <= 0 else "MOROSO"
        })

    df = pd.DataFrame(datos)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Cierre Anual {anio_actual}')
    
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Cierre_Anual_{anio_actual}_SiSE.xlsx"
    )

# --- REPORTE: MATRIZ MENSUAL EN EXCEL ---
@app.route('/admin/reporte-matriz-anual')
@login_required
def reporte_matriz_anual():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    anio_actual = datetime.now().year
    # Consultamos las casas y sus pagos del año actual
    casas = Casa.query.order_by(Casa.numero_casa).all()
    meses_nombres = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    matriz_datos = []
    for casa in casas:
        fila = {"Casa": casa.numero_casa, "Propietario": casa.dueno_nombre}
        
        # Obtenemos los meses que ya tienen pagos registrados
        pagos_anio = Pago.query.filter(
            Pago.casa_id == casa.id,
            func.extract('year', Pago.fecha) == anio_actual
        ).all()
        meses_pagados = [p.fecha.month for p in pagos_anio]

        for i in range(1, 13):
            fila[meses_nombres[i-1]] = "PAGADO" if i in meses_pagados else "PENDIENTE"
        
        fila["Saldo Total"] = casa.saldo_total
        matriz_datos.append(fila)

    df = pd.DataFrame(matriz_datos)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Escribimos los datos dejando espacio para el encabezado
        df.to_excel(writer, index=False, sheet_name='Matriz de Pagos', startrow=5)
        
        workbook = writer.book
        worksheet = writer.sheets['Matriz de Pagos']
        workbook.active = worksheet

        # Intentar insertar el Logo
        logo_path = os.path.join(app.root_path, 'static/img','logoEsperanza.png')
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                img.width, img.height = (80, 80) # Tamaño cuadrado pequeño
                worksheet.add_image(img, 'A1')
            except Exception as e:
                print(f"No se pudo cargar el logo en Excel: {e}")

        # Aplicar Estilos al Encabezado (Aquí es donde daba error)
        worksheet['C2'] = "CONJUNTO RESIDENCIAL LA ESPERANZA"
        worksheet['C2'].font = Font(size=16, bold=True, color="1E293B")
        
        worksheet['C3'] = f"MATRIZ GLOBAL DE PAGOS - AÑO {anio_actual}"
        worksheet['C3'].font = Font(size=12, bold=True)
        
        worksheet['C4'] = f"Reporte generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        worksheet['C4'].font = Font(size=10, italic=True)

    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"Matriz_Pagos_{anio_actual}.xlsx")


# --- REPORTE: RENDICIÓN DE CUENTAS EN PDF ---
@app.route('/admin/rendicion-cuentas-pdf')
@login_required
def rendicion_cuentas_pdf():
    if current_user.rol != 'admin':
        return redirect(url_for('inicio'))

    # Totales para el informe
    ing_ali = db.session.query(func.sum(Pago.monto)).scalar() or 0
    ing_ext = db.session.query(func.sum(IngresoExtra.monto)).scalar() or 0
    gastos = db.session.query(func.sum(Gasto.monto)).scalar() or 0
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    
    # Encabezado con Logo
    logo_path = os.path.join(app.root_path, 'static\img','logoEsperanza.png')
    if os.path.exists(logo_path):
        try:
            p.drawImage(logo_path, 20*mm, 260*mm, width=25*mm, preserveAspectRatio=True, mask='auto')
        except:
            pass

    p.setFont("Helvetica-Bold", 16)
    p.drawString(50*mm, 275*mm, "CONJUNTO RESIDENCIAL LA ESPERANZA")
    p.setFont("Helvetica", 10)
    p.drawString(50*mm, 270*mm, f"Informe Financiero de Gestión - Período {datetime.now().year}")
    p.line(20*mm, 255*mm, 190*mm, 255*mm)

    # Detalle de cuentas
    p.setFont("Helvetica-Bold", 12)
    p.drawString(25*mm, 240*mm, "RESUMEN DE CAJA:")
    
    p.setFont("Helvetica", 11)
    y = 230
    items = [
        (f"(+) Recaudación por Alícuotas:", ing_ali),
        (f"(+) Otros Ingresos / Extras:", ing_ext),
        (f"(-) Gastos de Mantenimiento:", gastos),
    ]
    
    for label, valor in items:
        p.drawString(30*mm, y*mm, label)
        p.drawRightString(180*mm, y*mm, f"${'{:,.2f}'.format(valor)}")
        y -= 8

    # Resultado Final
    p.line(140*mm, (y+4)*mm, 185*mm, (y+4)*mm)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(30*mm, (y-5)*mm, "SALDO NETO EN CAJA:")
    p.drawRightString(180*mm, (y-5)*mm, f"${'{:,.2f}'.format(ing_ali + ing_ext - gastos)}")

    p.showPage()
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="Rendicion_Cuentas.pdf", mimetype='application/pdf')



if __name__ == '__main__':
    with app.app_context():
        # Crea las tablas con la estructura limpia
        db.create_all()
        # Buscamos si ya existe el admin para no repetirlo
        admin_check = Usuario.query.filter_by(username='admin').first()
        
        if not admin_check:
            from werkzeug.security import generate_password_hash
            hashed_pw = generate_password_hash('admin123')
            
            # Creamos el único admin oficial
            admin_maestro = Usuario(
                username='admin',
                password=hashed_pw,
                rol='admin'
            )
            db.session.add(admin_maestro)
            db.session.commit()
            print("✅ SISTEMA LIMPIO: Admin creado (user: admin / pass: admin123)")
        else:
            print("ℹ️ El sistema ya tiene un administrador.")

    app.run(debug=True)